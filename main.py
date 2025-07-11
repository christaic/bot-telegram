import os, io
from datetime import datetime
from telegram import Update
from telegram.constants import MessageEntityType
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
from dotenv import load_dotenv

load_dotenv()
token = os.getenv("7717678907:aahidouqsn1tfueth-rrows5hgznpni8y50")

registro_estado = {}

def is_mentioning_bot(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    entities = update.message.entities or []
    for entity in entities:
        if entity.type == MessageEntityType.MENTION:
            text = update.message.text[entity.offset:entity.offset + entity.length]
            if text.lower() == f"@{context.bot.username.lower()}":
                return True
    return False

def is_reply_to_bot(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    return (
        update.message.reply_to_message and
        update.message.reply_to_message.from_user.id == context.bot.id
    )

def is_valid_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    chat_type = update.message.chat.type
    if chat_type == "private":
        return True
    return is_mentioning_bot(update, context) or is_reply_to_bot(update, context)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_valid_message(update, context):
        return
    group_id = update.effective_chat.id
    registro_estado[group_id] = {"step": 0, "data": {}}
    await update.message.reply_text("📍📝 hola, enviar nombre de calle y número de cuadra. ejemplo: av. los ingenieros - cuadra 8")

async def reiniciar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    group_id = update.effective_chat.id
    registro_estado[group_id] = {"step": 0, "data": {}}
    await update.message.reply_text("🔄 Hemos reiniciado el registro actual, usa /start para comenzar de nuevo.")

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_valid_message(update, context):
        return
    group_id = update.effective_chat.id
    if group_id not in registro_estado:
        await update.message.reply_text("❗ Usa /start para comenzar con el registro.")
        return
    mensaje = update.message.text.strip().lower()
    frases_consulta = [
        f"@{context.bot.username.lower()}",
        "¿en qué paso estoy?", "en qué paso estoy",
        "donde estoy", "dónde estoy",
        "me perdí", "me perdi",
        "paso actual", "sigo perdido",
        "bot?", "bot"
    ]
    if any(frase in mensaje for frase in frases_consulta):
        paso = registro_estado[group_id]["step"]
        mensajes_paso = {
            0: "📍 tranquilo 😊. estás en el paso 0: debes enviar el nombre de la calle y cuadra.",
            1: "🖼️ tranquilo 😊. estás en el paso 1: debes enviar la foto del antes.",
            2: "🖼️ no pasa nada 😊. estás en el paso 2: debes enviar la foto del después.",
            3: "🏷️ no pasa nada 😊. estás en el paso 3: debes enviar la foto de la etiqueta.",
            4: "📌 estoy para ayudarte 😊. estás en el paso 4: debes enviar la ubicación gps actual.",
        }
        await update.message.reply_text(mensajes_paso.get(paso, "😕 paso desconocido. usa /reiniciar y comencemos de nuevo"))
        return
    step = registro_estado[group_id]["step"]
    if step == 0:
        registro_estado[group_id]["data"]["calle"] = update.message.text
        registro_estado[group_id]["step"] = 1
        await update.message.reply_text("🖼️ 🖼️ conforme. ahora enviar la foto del antes. 🔔👀 recuerda que la foto se toma en vertical.")
    else:
        await update.message.reply_text("✅ sigue el flujo. envía lo que falta.")

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_valid_message(update, context):
        return
    group_id = update.effective_chat.id
    if group_id not in registro_estado:
        await update.message.reply_text("❗ Usa /start para comenzar con el registro.")
        return
    photo_file = await update.message.photo[-1].get_file()
    photo_bytes = await photo_file.download_as_bytearray()
    image = Image.open(io.BytesIO(photo_bytes))
    image_path = f"temp_{group_id}_{registro_estado[group_id]['step']}.jpg"
    image.save(image_path)
    step = registro_estado[group_id]["step"]
    if step == 1:
        registro_estado[group_id]["data"]["foto_antes"] = image_path
        registro_estado[group_id]["step"] = 2
        await update.message.reply_text("🖼️ conforme. ahora enviar la foto del despues. 🔔 recuerda que la foto se toma en vertical y debe ser del mismo angulo que la anterior.")
    elif step == 2:
        registro_estado[group_id]["data"]["foto_despues"] = image_path
        registro_estado[group_id]["step"] = 3
        await update.message.reply_text("🏷️ excelente. enviar foto de la etiqueta. 🔔 recuerda que la foto se toma en vertical.")
    elif step == 3:
        registro_estado[group_id]["data"]["foto_etiqueta"] = image_path
        registro_estado[group_id]["step"] = 4
        await update.message.reply_text("📌 genial, por ultimo enviar tu ubicación gps actual. 🔔 recuerda enviar las coordenadas desde el poste en el cual trabajo")
    else:
        await update.message.reply_text("✅ Fotos recibidas. Solo falta tu ubicación GPS actual.")

async def handle_location(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_valid_message(update, context):
        return
    group_id = update.effective_chat.id
    if group_id not in registro_estado:
        await update.message.reply_text("❗ Usa /start para comenzar con el registro.")
        return
    data = registro_estado[group_id]["data"]
    lat = update.message.location.latitude
    lon = update.message.location.longitude
    data["lat"] = lat
    data["lon"] = lon
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    filename = f"grupo_{group_id}_{fecha_actual}.xlsx"
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Fecha y hora", "Calle y cuadra", "Latitud", "Longitud", "Foto Antes", "Foto Después", "Foto Etiqueta"])
        for col in ['E', 'F', 'G']:
            ws.column_dimensions[col].width = 20
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), data["calle"], lat, lon, "", "", ""])
    row = ws.max_row
    for i, key in enumerate(["foto_antes", "foto_despues", "foto_etiqueta"]):
        img = ExcelImage(data[key])
        img.width, img.height = 120, 120
        col = chr(69 + i)
        ws.add_image(img, f"{col}{row}")
    ws.row_dimensions[row].height = 90
    wb.save(filename)
    registro_estado[group_id] = {"step": 0, "data": {}}
    await update.message.reply_text("✅ Registro guardado. Usa /start para registrar otro punto.")

async def exportar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_valid_message(update, context):
        return
    group_id = update.effective_chat.id
    fecha_actual = datetime.now().strftime("%Y-%m-%d")
    filename = f"grupo_{group_id}_{fecha_actual}.xlsx"
    if os.path.exists(filename):
        await update.message.reply_document(open(filename, "rb"))
    else:
        await update.message.reply_text("❌ No se encontró archivo para este grupo hoy.")

app = ApplicationBuilder().token("7717678907:aahidouqsn1tfueth-rrows5hgznpni8y50").build()
app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("reiniciar", reiniciar))
app.add_handler(CommandHandler("exportar", exportar))
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
app.add_handler(MessageHandler(filters.LOCATION, handle_location))
app.run_polling()